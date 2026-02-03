"""
Extraction module for processing Excel files and merging them into a single output file.
"""
import os
import re
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference


def extract_semester_name(cell_value):
    """
    Extract semester name from cell A1.
    Expected format: "2025-2026 ősz" félév - téma osztályzatok
    Returns: 2025-2026 ősz
    """
    if not cell_value:
        return None
    
    # Extract text between quotes
    match = re.search(r'"([^"]+)"', str(cell_value))
    if match:
        return match.group(1)
    
    return None


def load_course_aliases(script_dir):
    """
    Load course aliases from CourseAliases.xlsx.
    Returns a dictionary mapping CourseNeptunCode to (CourseAlias, CoursesSemesterIndex).
    """
    aliases_path = os.path.join(script_dir, "CourseAliases.xlsx")
    
    if not os.path.exists(aliases_path):
        raise Exception(f"CourseAliases.xlsx not found at {aliases_path}")
    
    print("Loading course aliases...")
    
    wb = load_workbook(aliases_path, data_only=True)
    ws = wb.active
    
    aliases = {}
    # Skip header row, start from row 2
    for row_idx in range(2, ws.max_row + 1):
        neptun_code = ws.cell(row_idx, 2).value  # Column B: CourseNeptunCode
        alias = ws.cell(row_idx, 3).value         # Column C: CourseAlias
        semester_index = ws.cell(row_idx, 4).value  # Column D: CoursesSemesterIndex
        
        if neptun_code and alias:
            aliases[str(neptun_code).strip()] = {
                'alias': str(alias).strip(),
                'semester_index': semester_index if semester_index is not None else ''
            }
    
    wb.close()
    print(f"✓ Loaded {len(aliases)} course aliases")
    
    return aliases


def is_valid_row(row_data):
    """
    Check if a row should be included in the output.
    Skip rows where Grade or GradedBy is missing or contains only "-"
    """
    grade = row_data.get('Grade', '')
    graded_by = row_data.get('GradedBy', '')
    
    # Convert to string and strip
    grade_str = str(grade).strip() if grade is not None else ''
    graded_by_str = str(graded_by).strip() if graded_by is not None else ''
    
    # Skip if either is empty or just a dash
    if not grade_str or grade_str == '-':
        return False
    if not graded_by_str or graded_by_str == '-':
        return False
    
    return True


def process_excel_file(file_path, course_aliases):
    """
    Extract data from a single Excel file.
    
    Args:
        file_path: Path to the Excel file
        course_aliases: Dictionary mapping CourseNeptunCode to CourseAlias
    
    Returns:
        List of dictionaries containing extracted data
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Extract semester ID from filename (e.g., PortalResults_SemesterId05.xlsx -> 5)
        filename = os.path.basename(file_path)
        semester_id_match = re.search(r'SemesterId(\d+)', filename)
        semester_id = int(semester_id_match.group(1)) if semester_id_match else 0
        
        # Extract semester name from A1
        semester_name = extract_semester_name(ws['A1'].value)
        if not semester_name:
            print(f"  ⚠ Warning: Could not extract semester name from {file_path}")
            semester_name = "Unknown"
        
        # Table starts at row 4 (header row)
        # We need columns B, C, D, E, F (indices 2, 3, 4, 5, 6)
        data_rows = []
        missing_aliases = set()
        
        # Start from row 5 (data rows after header)
        for row_idx in range(5, ws.max_row + 1):
            # Check if row is empty (stop at first empty row)
            row_cells = [ws.cell(row_idx, col_idx).value for col_idx in range(2, 7)]
            
            # If all cells are None or empty, stop processing
            if all(cell is None or str(cell).strip() == '' for cell in row_cells):
                break
            
            # Extract data from columns B through F
            course_neptun_code = ws.cell(row_idx, 4).value  # Column D
            
            row_data = {
                'SemesterName': semester_name,
                'SemesterId': semester_id,
                'StudentNeptunCode': ws.cell(row_idx, 2).value,  # Column B
                'CourseName': ws.cell(row_idx, 3).value,          # Column C
                'CourseNeptunCode': course_neptun_code,            # Column D
                'Grade': ws.cell(row_idx, 5).value,               # Column E
                'GradedBy': ws.cell(row_idx, 6).value             # Column F
            }
            
            # Only add row if it passes validation
            if is_valid_row(row_data):
                # Lookup CourseAlias and CoursesSemesterIndex
                neptun_key = str(course_neptun_code).strip() if course_neptun_code else ""
                if neptun_key in course_aliases:
                    alias = course_aliases[neptun_key]['alias']
                    semester_index = course_aliases[neptun_key]['semester_index']
                    
                    # Skip rows where CourseAlias is "NA"
                    if alias == "NA":
                        continue
                    
                    row_data['CourseAlias'] = alias
                    row_data['CoursesSemesterIndex'] = semester_index
                else:
                    missing_aliases.add(neptun_key)
                    row_data['CourseAlias'] = None
                    row_data['CoursesSemesterIndex'] = None
                
                data_rows.append(row_data)
        
        wb.close()
        
        # Check for missing aliases
        if missing_aliases:
            raise Exception(f"Missing course aliases for Neptun codes: {', '.join(sorted(missing_aliases))}")
        
        return data_rows
        
    except Exception as e:
        raise Exception(f"Error processing file {file_path}: {e}")


def create_statistics_matrix(all_data, output_wb):
    """
    Create statistics matrix worksheet with headcounts and chart.
    
    Args:
        all_data: List of data dictionaries
        output_wb: Output workbook object
    """
    print("\nCreating statistics matrix...")
    
    # Create new worksheet
    stats_ws = output_wb.create_sheet("CourseHeadCounts")
    
    # Collect unique semesters and course semester indices
    semesters = []
    semester_set = set()
    for row in all_data:
        semester = row['SemesterName']
        if semester not in semester_set:
            semesters.append(semester)
            semester_set.add(semester)
    
    course_semester_indices = sorted(set(row['CoursesSemesterIndex'] for row in all_data))
    
    # Count occurrences
    counts = defaultdict(lambda: defaultdict(int))
    for row in all_data:
        semester = row['SemesterName']
        semester_index = row['CoursesSemesterIndex']
        counts[semester][semester_index] += 1
    
    # Write header row
    stats_ws.cell(1, 1, "SemesterName")
    for col_idx, semester_index in enumerate(course_semester_indices, 2):
        stats_ws.cell(1, col_idx, semester_index)
    
    # Write data rows
    for row_idx, semester in enumerate(semesters, 2):
        stats_ws.cell(row_idx, 1, semester)
        for col_idx, semester_index in enumerate(course_semester_indices, 2):
            count = counts[semester][semester_index]
            stats_ws.cell(row_idx, col_idx, count)
    
    # Create stacked column chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Evolution of headcounts"
    chart.y_axis.title = "Student Count"
    chart.x_axis.title = "Semester"
    chart.grouping = "stacked"
    chart.overlap = 100
    
    # Data for chart (all course semester indices)
    data = Reference(stats_ws, min_col=2, min_row=1, max_col=len(course_semester_indices) + 1, max_row=len(semesters) + 1)
    # Categories (semester names)
    cats = Reference(stats_ws, min_col=1, min_row=2, max_row=len(semesters) + 1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Position chart below the data
    chart_row = len(semesters) + 3
    stats_ws.add_chart(chart, f"A{chart_row}")
    
    print(f"✓ Created statistics matrix: {len(semesters)} semesters × {len(course_semester_indices)} courses")
    print(f"✓ Added stacked column chart")


def create_population_timelines(all_data, output_wb):
    """
    Create population timelines worksheet with shifted data.
    
    Args:
        all_data: List of data dictionaries
        output_wb: Output workbook object
    """
    print("\nCreating population timelines...")
    
    # Create new worksheet
    timeline_ws = output_wb.create_sheet("PopulationTimelines")
    
    # Collect unique semester IDs and course semester indices
    semester_ids = []
    semester_id_set = set()
    for row in all_data:
        semester_id = row['SemesterId']
        if semester_id not in semester_id_set:
            semester_ids.append(semester_id)
            semester_id_set.add(semester_id)
    
    semester_ids.sort()
    max_semester_id = max(semester_ids)
    course_semester_indices = sorted(set(row['CoursesSemesterIndex'] for row in all_data))
    
    # Count occurrences by SemesterId
    counts = defaultdict(lambda: defaultdict(int))
    for row in all_data:
        semester_id = row['SemesterId']
        semester_index = row['CoursesSemesterIndex']
        counts[semester_id][semester_index] += 1
    
    # Write header row
    # Header starts with SemesterId, then course semester indices (repeated to accommodate shifts)
    timeline_ws.cell(1, 1, "SemesterId")
    # We need enough columns for the shifted data
    # Max shift is (max_semester_id - min_semester_id)
    # Total columns needed: 1 (SemesterId) + max_shift + len(course_semester_indices)
    max_shift = max_semester_id - min(semester_ids)
    total_data_cols = max_shift + len(course_semester_indices)
    
    # Write course semester indices as headers (they appear multiple times due to shifts)
    for col_idx in range(total_data_cols):
        # Determine which course semester index this column represents
        course_idx = col_idx % len(course_semester_indices)
        if col_idx < len(course_semester_indices):
            timeline_ws.cell(1, col_idx + 2, course_semester_indices[course_idx])
    
    # Actually, let's just put all the course semester indices in order, starting from column 2
    for col_idx, semester_index in enumerate(course_semester_indices, 0):
        timeline_ws.cell(1, col_idx + 2, semester_index)
    
    # Write data rows with shifts
    for row_idx, semester_id in enumerate(semester_ids, 2):
        # Write SemesterId in column 1
        timeline_ws.cell(row_idx, 1, semester_id)
        
        # Calculate shift: N = max_semester_id - current_semester_id
        shift = max_semester_id - semester_id
        
        # Write data starting from column (2 + shift)
        for col_offset, semester_index in enumerate(course_semester_indices):
            count = counts[semester_id][semester_index]
            target_col = 2 + shift + col_offset
            timeline_ws.cell(row_idx, target_col, count)
    
    print(f"✓ Created population timelines: {len(semester_ids)} semesters with shifts applied")


def merge_all_files(downloads_dir, output_dir):
    """
    Process all Excel files in downloads directory and merge into single output file.
    
    Args:
        downloads_dir: Directory containing downloaded Excel files
        output_dir: Directory where output file will be saved
    
    Returns:
        Path to the output file
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Load course aliases
    course_aliases = load_course_aliases(script_dir)
    
    # Get all Excel files from downloads directory
    excel_files = [f for f in os.listdir(downloads_dir) 
                   if f.endswith(('.xlsx', '.xls')) and os.path.isfile(os.path.join(downloads_dir, f))]
    
    if not excel_files:
        raise Exception(f"No Excel files found in {downloads_dir}")
    
    print(f"\nProcessing {len(excel_files)} Excel files...")
    
    # Collect all data
    all_data = []
    
    for idx, filename in enumerate(excel_files, 1):
        file_path = os.path.join(downloads_dir, filename)
        print(f"Processing file {idx}/{len(excel_files)}: {filename}...", end=" ", flush=True)
        
        try:
            data_rows = process_excel_file(file_path, course_aliases)
            all_data.extend(data_rows)
            print(f"✓ Extracted {len(data_rows)} rows")
        except Exception as e:
            print(f"✗ Error: {e}")
            raise
    
    # Create output workbook
    print(f"\nCreating output file with {len(all_data)} total rows...")
    
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "AllSemesterData"
    
    # Write header
    headers = ['SemesterName', 'SemesterId', 'StudentNeptunCode', 'CourseName', 'CourseNeptunCode', 'Grade', 'GradedBy', 'CourseAlias', 'CoursesSemesterIndex']
    for col_idx, header in enumerate(headers, 1):
        output_ws.cell(1, col_idx, header)
    
    # Sort data by SemesterId
    all_data.sort(key=lambda x: x['SemesterId'])
    
    # Write data rows
    for row_idx, data_row in enumerate(all_data, 2):
        output_ws.cell(row_idx, 1, data_row['SemesterName'])
        output_ws.cell(row_idx, 2, data_row['SemesterId'])
        output_ws.cell(row_idx, 3, data_row['StudentNeptunCode'])
        output_ws.cell(row_idx, 4, data_row['CourseName'])
        output_ws.cell(row_idx, 5, data_row['CourseNeptunCode'])
        output_ws.cell(row_idx, 6, data_row['Grade'])
        output_ws.cell(row_idx, 7, data_row['GradedBy'])
        output_ws.cell(row_idx, 8, data_row['CourseAlias'])
        output_ws.cell(row_idx, 9, data_row['CoursesSemesterIndex'])
    
    # Create statistics matrix
    create_statistics_matrix(all_data, output_wb)
    
    # Create population timelines
    create_population_timelines(all_data, output_wb)
    
    # Save output file
    output_path = os.path.join(output_dir, "AllSemesterProjectStats.xlsx")
    output_wb.save(output_path)
    output_wb.close()
    
    print(f"✓ Output file saved: {output_path}")
    return output_path


if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    downloads_dir = os.path.join(script_dir, "downloads")
    output_dir = os.path.join(script_dir, "output")
    
    try:
        merge_all_files(downloads_dir, output_dir)
    except Exception as e:
        print(f"Error: {e}")
        exit(1)
