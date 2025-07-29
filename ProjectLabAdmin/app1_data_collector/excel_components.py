"""
Excel formatting and styling utilities for MKXLSX
Separated according to SRP
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Any
import logging


class ExcelStyler:
    """Responsible for Excel styling and formatting"""
    
    def __init__(self):
        # Define standard colors and styles
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.title_font = Font(bold=True, size=14)
        self.section_font = Font(bold=True, size=12)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def style_header_cell(self, cell):
        """Apply header styling to a cell"""
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.border = self.border
        cell.alignment = self.center_alignment
    
    def style_title_cell(self, cell):
        """Apply title styling to a cell"""
        cell.font = self.title_font
        cell.alignment = self.center_alignment
    
    def style_section_title_cell(self, cell):
        """Apply section title styling to a cell"""
        cell.font = self.section_font
    
    def style_data_cell(self, cell):
        """Apply data cell styling"""
        cell.border = self.border
        cell.alignment = self.center_alignment
    
    def style_bold_cell(self, cell):
        """Apply bold styling to a cell"""
        cell.font = Font(bold=True)


class ExcelTableBuilder:
    """Responsible for building Excel table structures"""
    
    def __init__(self, worksheet, styler: ExcelStyler, logger: logging.Logger):
        self.worksheet = worksheet
        self.styler = styler
        self.logger = logger
    
    def create_header_row(self, row: int, headers: list, start_col: int = 1) -> None:
        """
        Create a styled header row
        
        Args:
            row: Row number
            headers: List of header texts
            start_col: Starting column number
        """
        for col, header in enumerate(headers, start_col):
            cell = self.worksheet.cell(row=row, column=col, value=header)
            self.styler.style_header_cell(cell)
    
    def create_data_rows(self, start_row: int, data: list, start_col: int = 1) -> int:
        """
        Create data rows in the table
        
        Args:
            start_row: Starting row number
            data: List of row data (each item is a list/tuple)
            start_col: Starting column number
            
        Returns:
            Next available row number
        """
        current_row = start_row
        
        for row_data in data:
            for col, value in enumerate(row_data, start_col):
                cell = self.worksheet.cell(row=current_row, column=col, value=value)
                self.styler.style_data_cell(cell)
            current_row += 1
        
        return current_row
    
    def set_column_widths(self, column_widths: dict) -> None:
        """
        Set column widths
        
        Args:
            column_widths: Dict mapping column letters to widths
        """
        for col_letter, width in column_widths.items():
            self.worksheet.column_dimensions[col_letter].width = width


class PlannerTableCreator:
    """Responsible for creating the planner table"""
    
    def __init__(self, worksheet, styler: ExcelStyler, table_builder: ExcelTableBuilder, logger: logging.Logger):
        self.worksheet = worksheet
        self.styler = styler
        self.table_builder = table_builder
        self.logger = logger
        self.time_slots = ["8:00-10:00", "10:00-12:00", "12:00-14:00", "14:00-16:00"]
        self.weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        self.rooms = ["QB203", "QBF14", "QBF15"]
    
    def create_planner_table(self, start_row: int = 3) -> int:
        """
        Create the main planner table
        
        Args:
            start_row: Starting row for the planner table
            
        Returns:
            Next available row after the table
        """
        try:
            self.logger.info("Creating planner table...")
            
            # Add title
            self.worksheet.cell(row=start_row-1, column=1, value="SESSION PLANNER")
            self.styler.style_title_cell(self.worksheet.cell(row=start_row-1, column=1))
            
            # Create column headers (weekday-room combinations)
            headers = ["Time Slot"]
            for weekday in self.weekdays:
                for room in self.rooms:
                    headers.append(f"{weekday}\n{room}")
            
            self.table_builder.create_header_row(start_row, headers)
            
            # Set column widths
            column_widths = {'A': 15}
            for col in range(2, len(headers) + 1):
                column_widths[get_column_letter(col)] = 12
            
            self.table_builder.set_column_widths(column_widths)
            
            # Create data rows for time slots
            current_row = start_row + 1
            for time_slot in self.time_slots:
                cell = self.worksheet.cell(row=current_row, column=1, value=time_slot)
                self.styler.style_data_cell(cell)
                
                # Create empty cells for session entries
                for col in range(2, len(headers) + 1):
                    cell = self.worksheet.cell(row=current_row, column=col, value="")
                    self.styler.style_data_cell(cell)
                
                current_row += 1
            
            self.logger.info(f"Planner table created from row {start_row} to {current_row - 1}")
            return current_row + 1  # Leave a gap after the table
            
        except Exception as e:
            self.logger.error(f"Error creating planner table: {e}")
            return start_row


class StatisticsCalculator:
    """Responsible for calculating session planning statistics"""
    
    def __init__(self, logger: logging.Logger, statistics_course_codes: set):
        self.logger = logger
        self.statistics_course_codes = statistics_course_codes
    
    def calculate_topic_statistics(self, fused_data: dict) -> dict:
        """
        Calculate topic category statistics from fused data
        
        Args:
            fused_data: The fused student data dictionary
            
        Returns:
            Dictionary containing statistics
        """
        hungarian_topic_categories = {}
        english_student_count = 0
        
        if fused_data and 'students' in fused_data:
            for student in fused_data['students']:
                # Check if student is enrolled in any of the statistics course codes
                enrolled_courses = student.get('enrolled_courses', [])
                has_relevant_course = any(
                    course.get('course_code') in self.statistics_course_codes 
                    for course in enrolled_courses
                )
                
                # Only count students who are enrolled in relevant courses and have topics
                if has_relevant_course and student.get('has_topic') and student.get('topic_category'):
                    # Check if this is an English topic (Z-ENG prefix)
                    is_english_topic = student.get('is_english_topic', False)
                    
                    if is_english_topic:
                        # Count English topics separately
                        english_student_count += 1
                    else:
                        # Count Hungarian topics by category
                        category = student['topic_category']
                        hungarian_topic_categories[category] = hungarian_topic_categories.get(category, 0) + 1
        
        return {
            'hungarian_topic_categories': hungarian_topic_categories,
            'english_student_count': english_student_count,
            'total_hungarian_students': sum(hungarian_topic_categories.values()),
            'total_students': sum(hungarian_topic_categories.values()) + english_student_count
        }
    
    def calculate_required_sessions(self, student_count: int, students_per_session: int = 9) -> int:
        """Calculate the number of sessions required for a given number of students"""
        import math
        return math.ceil(student_count / students_per_session) if student_count > 0 else 0


class StatisticsTableCreator:
    """Responsible for creating statistics tables"""
    
    def __init__(self, worksheet, styler: ExcelStyler, table_builder: ExcelTableBuilder, 
                 stats_calculator: StatisticsCalculator, logger: logging.Logger):
        self.worksheet = worksheet
        self.styler = styler
        self.table_builder = table_builder
        self.stats_calculator = stats_calculator
        self.logger = logger
        self.session_types = ["SW", "HW", "ENG", "GEN", "ONLINE", "SPARE", "NONE", "ROBONAUT", "AI"]
    
    def create_statistics_section(self, start_row: int, fused_data: dict) -> int:
        """
        Create the statistics section with topic and session type statistics
        
        Args:
            start_row: Starting row for statistics section
            fused_data: The fused student data
            
        Returns:
            Next available row after statistics section
        """
        try:
            self.logger.info("Creating statistics section...")
            current_row = start_row
            
            # Section 1: Topic Category Statistics
            current_row = self._create_topic_statistics_table(current_row, fused_data)
            current_row += 2  # Gap between sections
            
            # Section 2: Session Type Statistics
            current_row = self._create_session_type_statistics_table(current_row)
            
            self.logger.info(f"Statistics section created from row {start_row} to {current_row}")
            return current_row
            
        except Exception as e:
            self.logger.error(f"Error creating statistics section: {e}")
            return start_row
    
    def _create_topic_statistics_table(self, start_row: int, fused_data: dict) -> int:
        """Create the topic category statistics table"""
        current_row = start_row
        
        # Section title
        self.worksheet.cell(row=current_row, column=1, value="TOPIC CATEGORY STATISTICS")
        self.styler.style_section_title_cell(self.worksheet.cell(row=current_row, column=1))
        current_row += 1
        
        # Headers
        headers = ["Category", "Students", "Required Sessions (9 per session)"]
        self.table_builder.create_header_row(current_row, headers)
        current_row += 1
        
        # Calculate statistics
        stats = self.stats_calculator.calculate_topic_statistics(fused_data)
        
        # Add rows for each Hungarian topic category
        for category, count in stats['hungarian_topic_categories'].items():
            required_sessions = self.stats_calculator.calculate_required_sessions(count)
            
            self.worksheet.cell(row=current_row, column=1, value=category)
            self.worksheet.cell(row=current_row, column=2, value=count)
            self.worksheet.cell(row=current_row, column=3, value=required_sessions)
            current_row += 1
        
        # Add separate row for English topics (ENG category)
        if stats['english_student_count'] > 0:
            required_sessions = self.stats_calculator.calculate_required_sessions(stats['english_student_count'])
            self.worksheet.cell(row=current_row, column=1, value="ENG")
            self.worksheet.cell(row=current_row, column=2, value=stats['english_student_count'])
            self.worksheet.cell(row=current_row, column=3, value=required_sessions)
            current_row += 1
        
        # Add total row
        total_students = stats['total_students']
        total_sessions = self.stats_calculator.calculate_required_sessions(total_students)
        
        self.worksheet.cell(row=current_row, column=1, value="TOTAL")
        self.styler.style_bold_cell(self.worksheet.cell(row=current_row, column=1))
        self.worksheet.cell(row=current_row, column=2, value=total_students)
        self.styler.style_bold_cell(self.worksheet.cell(row=current_row, column=2))
        self.worksheet.cell(row=current_row, column=3, value=total_sessions)
        self.styler.style_bold_cell(self.worksheet.cell(row=current_row, column=3))
        current_row += 1
        
        return current_row
    
    def _create_session_type_statistics_table(self, start_row: int) -> int:
        """Create the session type statistics table with Excel formulas"""
        current_row = start_row
        
        # Section title
        self.worksheet.cell(row=current_row, column=1, value="SESSION TYPE STATISTICS")
        self.styler.style_section_title_cell(self.worksheet.cell(row=current_row, column=1))
        current_row += 1
        
        # Headers
        headers = ["Session Type", "Count", "Description"]
        self.table_builder.create_header_row(current_row, headers)
        current_row += 1
        
        # Session type descriptions
        descriptions = {
            "SW": "Software session",
            "HW": "Hardware and software session",
            "ENG": "English course presentation",
            "GEN": "Generic session for all types",
            "ONLINE": "Online generic session",
            "SPARE": "Spare time slot for later use",
            "NONE": "Not used timeslot",
            "ROBONAUT": "Special ROBONAUT session",
            "AI": "Special AI session"
        }
        
        # Add formula rows for each session type
        for session_type in self.session_types:
            # Create COUNTIF formula to count occurrences in planner table
            # The planner table is in range B4:P7 (columns B to P, rows 4 to 7)
            formula = f'=COUNTIF(B4:P7,"{session_type}")'
            description = descriptions.get(session_type, f"{session_type} session")
            
            self.worksheet.cell(row=current_row, column=1, value=session_type)
            self.worksheet.cell(row=current_row, column=2, value=formula)
            self.worksheet.cell(row=current_row, column=3, value=description)
            current_row += 1
        
        # Set column widths for statistics section
        column_widths = {'A': 20, 'B': 15, 'C': 30}
        self.table_builder.set_column_widths(column_widths)
        
        return current_row
