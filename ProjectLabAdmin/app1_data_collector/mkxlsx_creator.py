"""
MKXLSX - Excel table creation for session planning
This module creates an Excel table using DLFUSION data for session schedule planning.
"""

import json
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math
import sys

sys.path.append(str(Path(__file__).parent.parent))
from shared import setup_logging


class SessionPlannerExcelCreator:
    """Creator for session planning Excel table"""
    
    def __init__(self):
        """Initialize the Excel creator"""
        self.logger = setup_logging(__name__)
        self.fused_data = None
        self.workbook = None
        self.worksheet = None
        
        # Define time slots and rooms from specification
        self.time_slots = [
            "8:00-10:00",
            "10:00-12:00", 
            "12:00-14:00",
            "14:00-16:00"
        ]
        
        self.weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        self.rooms = ["QB203", "QBF14", "QBF15"]
        
        # Course codes to consider for statistics (from specification)
        # Updated to include courses where students with topics are actually enrolled
        self.STATISTICS_COURSE_CODES = {
            'BMEVIAUMT00', 'BMEVIAUMT10', 'BMEVIAUMT12', 'BMEVIAUM026', 
            'BMEVIAUAL01', 'BMEVIAUAL03', 'BMEVIAUAL04', 'BMEVIAUAL05', 
            'BMEVIAUML10', 'BMEVIAUML12', 'BMEVIAUML11', 'BMEVIAUML13', 
            'BMEVIAUM039', 'BMEVIAUAT02', 'BMEVIAUMT11'
        }
        
        # Session types from specification
        self.session_types = ["SW", "HW", "ENG", "GEN", "ONLINE", "SPARE", "NONE"]
        
        # Additional session types mentioned in specification
        self.special_session_types = ["ROBONAUT", "AI"]
        
        self.all_session_types = self.session_types + self.special_session_types
    
    def load_fused_data(self, json_file: Optional[Path] = None) -> bool:
        """
        Load fused data from JSON file
        
        Args:
            json_file: Path to fused JSON file (defaults to data/fused_student_data.json)
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if json_file is None:
                json_file = Path(__file__).parent.parent / "data" / "fused_student_data.json"
            
            if not json_file.exists():
                self.logger.error(f"Fused data file not found: {json_file}")
                return False
            
            self.logger.info(f"Loading fused data from: {json_file}")
            
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            self.fused_data = data
            self.logger.info(f"Loaded fused data with {len(data['students'])} students")
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading fused data: {e}")
            return False
    
    def create_planner_table(self, start_row: int = 3) -> int:
        """
        Create the planner table section
        
        Args:
            start_row: Starting row for the planner table
            
        Returns:
            Next available row after the table
        """
        try:
            self.logger.info("Creating planner table...")
            
            # Title for planner section
            self.worksheet.cell(row=start_row-1, column=1, value="SESSION PLANNER")
            self.worksheet.cell(row=start_row-1, column=1).font = Font(bold=True, size=14)
            
            # Create column headers (weekday-room pairs)
            col = 2  # Start from column B (leave column A for time slots)
            for weekday in self.weekdays:
                for room in self.rooms:
                    header = f"{weekday}\n{room}"
                    cell = self.worksheet.cell(row=start_row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                    
                    # Set column width
                    self.worksheet.column_dimensions[get_column_letter(col)].width = 12
                    col += 1
            
            # Create row headers (time slots) and empty cells for planning
            current_row = start_row + 1
            for time_slot in self.time_slots:
                # Time slot header
                cell = self.worksheet.cell(row=current_row, column=1, value=time_slot)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                
                # Empty cells for user input (session types)
                for col in range(2, 2 + len(self.weekdays) * len(self.rooms)):
                    cell = self.worksheet.cell(row=current_row, column=col, value="")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Add border
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
                
                current_row += 1
            
            # Set width for time slot column
            self.worksheet.column_dimensions['A'].width = 15
            
            self.logger.info(f"Planner table created from row {start_row} to {current_row-1}")
            return current_row + 2  # Leave some space before statistics
            
        except Exception as e:
            self.logger.error(f"Error creating planner table: {e}")
            return start_row
    
    def create_statistics_section(self, start_row: int) -> int:
        """
        Create the statistics section
        
        Args:
            start_row: Starting row for the statistics section
            
        Returns:
            Next available row after the statistics
        """
        try:
            self.logger.info("Creating statistics section...")
            
            current_row = start_row
            
            # Section 1: Topic Category Statistics
            self.worksheet.cell(row=current_row, column=1, value="TOPIC CATEGORY STATISTICS")
            self.worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 2
            
            # Headers for topic statistics
            headers = ["Category", "Students", "Required Sessions (9 per session)"]
            for col, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            current_row += 1
            
            # Calculate topic category statistics from fused data
            # Include all students with topics, regardless of course enrollment
            hungarian_topic_categories = {}
            english_student_count = 0
            
            if self.fused_data and 'students' in self.fused_data:
                for student in self.fused_data['students']:
                    # Count all students who have topics
                    if student.get('has_topic') and student.get('topic_category'):
                        # Check if this is an English topic (Z-ENG prefix)
                        is_english_topic = student.get('is_english_topic', False)
                        
                        if is_english_topic:
                            # Count English topics separately
                            english_student_count += 1
                        else:
                            # Count Hungarian topics by category
                            category = student['topic_category']
                            hungarian_topic_categories[category] = hungarian_topic_categories.get(category, 0) + 1
            
            self.logger.info(f"Hungarian categories: {hungarian_topic_categories}")
            self.logger.info(f"English student count: {english_student_count}")
            
            # Add rows for each Hungarian topic category
            for category, count in hungarian_topic_categories.items():
                required_sessions = math.ceil(count / 9)  # 9 students per session
                
                self.worksheet.cell(row=current_row, column=1, value=category)
                self.worksheet.cell(row=current_row, column=2, value=count)
                self.worksheet.cell(row=current_row, column=3, value=required_sessions)
                current_row += 1
            
            # Add separate row for English topics (ENG category)
            if english_student_count > 0:
                required_sessions = math.ceil(english_student_count / 9)
                self.worksheet.cell(row=current_row, column=1, value="ENG")
                self.worksheet.cell(row=current_row, column=2, value=english_student_count)
                self.worksheet.cell(row=current_row, column=3, value=required_sessions)
                current_row += 1
            
            # Add total row (Hungarian + English)
            total_hungarian_students = sum(hungarian_topic_categories.values())
            total_students = total_hungarian_students + english_student_count
            total_sessions = math.ceil(total_students / 9)
            
            self.worksheet.cell(row=current_row, column=1, value="TOTAL")
            self.worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            self.worksheet.cell(row=current_row, column=2, value=total_students)
            self.worksheet.cell(row=current_row, column=2).font = Font(bold=True)
            self.worksheet.cell(row=current_row, column=3, value=total_sessions)
            self.worksheet.cell(row=current_row, column=3).font = Font(bold=True)
            
            current_row += 3
            
            # Section 2: Session Type Statistics (with Excel formulas)
            self.worksheet.cell(row=current_row, column=1, value="SESSION TYPE STATISTICS")
            self.worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 2
            
            # Headers for session type statistics
            headers = ["Session Type", "Count in Planner", "Description"]
            for col, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True) 
                cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
            
            current_row += 1
            
            # Session type descriptions
            session_descriptions = {
                "SW": "Software session",
                "HW": "Hardware and software session", 
                "ENG": "English course presentation",
                "GEN": "Generic session for all types",
                "ONLINE": "Online generic session",
                "SPARE": "Spare time slot for later use",
                "NONE": "Not used timeslot",
                "ROBONAUT": "Special ROBONAUT session",
                "AI": "Special AI session"
            }
            
            # Calculate the range for the planner table (where user will input session types)
            planner_start_row = 4  # First data row of planner
            planner_end_row = planner_start_row + len(self.time_slots) - 1
            planner_start_col = 2  # Column B
            planner_end_col = planner_start_col + len(self.weekdays) * len(self.rooms) - 1
            
            planner_range = f"{get_column_letter(planner_start_col)}{planner_start_row}:{get_column_letter(planner_end_col)}{planner_end_row}"
            
            # Add rows for each session type with Excel COUNTIF formulas
            for session_type in self.all_session_types:
                self.worksheet.cell(row=current_row, column=1, value=session_type)
                
                # Excel formula to count occurrences in planner table
                formula = f'=COUNTIF({planner_range},"{session_type}")'
                self.worksheet.cell(row=current_row, column=2, value=formula)
                
                description = session_descriptions.get(session_type, "Custom session type")
                self.worksheet.cell(row=current_row, column=3, value=description)
                
                current_row += 1
            
            # Set column widths for statistics section
            self.worksheet.column_dimensions['A'].width = 20
            self.worksheet.column_dimensions['B'].width = 15
            self.worksheet.column_dimensions['C'].width = 30
            
            self.logger.info(f"Statistics section created from row {start_row} to {current_row-1}")
            return current_row
            
        except Exception as e:
            self.logger.error(f"Error creating statistics section: {e}")
            return start_row
    
    def add_instructions_and_formatting(self):
        """Add instructions and final formatting to the Excel file"""
        try:
            # Add title and instructions at the top
            self.worksheet.cell(row=1, column=1, value="PROJECT LABORATORY SESSION PLANNER")
            self.worksheet.cell(row=1, column=1).font = Font(bold=True, size=16, color="2F4F4F")
            
            # Instructions
            instructions = (
                "Instructions: Fill in the planner table with session type codes (SW, HW, ENG, GEN, ONLINE, SPARE, NONE, or custom codes like ROBONAUT, AI). "
                "Statistics will update automatically."
            )
            self.worksheet.cell(row=2, column=1, value=instructions)
            self.worksheet.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
            self.worksheet.cell(row=2, column=1).font = Font(italic=True, size=10)
            
            # Merge cells for title and instructions
            self.worksheet.merge_cells(f'A1:{get_column_letter(len(self.weekdays) * len(self.rooms) + 1)}1')
            self.worksheet.merge_cells(f'A2:{get_column_letter(len(self.weekdays) * len(self.rooms) + 1)}2')
            
            # Set row heights
            self.worksheet.row_dimensions[1].height = 25
            self.worksheet.row_dimensions[2].height = 40
            
            self.logger.info("Added instructions and formatting")
            
        except Exception as e:
            self.logger.error(f"Error adding instructions and formatting: {e}")
    
    def create_session_planning_excel(self, output_file: Optional[Path] = None) -> bool:
        """
        Create the complete session planning Excel file
        
        Args:
            output_file: Path for output Excel file (defaults to data/session_planner.xlsx)
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if output_file is None:
                output_file = Path(__file__).parent.parent / "data" / "session_planner.xlsx"
            
            # Ensure output directory exists
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            self.logger.info(f"Creating session planning Excel file: {output_file}")
            
            # Load fused data first
            if not self.load_fused_data():
                self.logger.error("Failed to load fused data")
                return False
            
            # Create workbook and worksheet
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Session Planner"
            
            # Add instructions and title
            self.add_instructions_and_formatting()
            
            # Create the planner table
            next_row = self.create_planner_table(start_row=3)
            
            # Create the statistics section
            self.create_statistics_section(start_row=next_row)
            
            # Save the workbook  
            self.workbook.save(output_file)
            
            self.logger.info(f"Session planning Excel file created: {output_file}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating session planning Excel file: {e}")
            return False
    
    def get_planning_summary(self) -> Dict[str, Any]:
        """
        Get a summary of the planning data
        
        Returns:
            Dictionary with planning statistics
        """
        if not self.fused_data:
            return {'error': 'No fused data available'}
        
        # Calculate topic category statistics (separate Hungarian and English)
        # Only consider students enrolled in the specified statistics course codes
        hungarian_topic_categories = {}
        english_student_count = 0
        students_with_topics = 0
        total_students = len(self.fused_data.get('students', []))
        
        for student in self.fused_data.get('students', []):
            # Check if student is enrolled in any of the statistics course codes
            enrolled_courses = student.get('enrolled_courses', [])
            has_relevant_course = any(
                course.get('course_code') in self.STATISTICS_COURSE_CODES 
                for course in enrolled_courses
            )
            
            # Only count students who are enrolled in relevant courses and have topics
            if has_relevant_course and student.get('has_topic') and student.get('topic_category'):
                students_with_topics += 1
                # Check if this is an English topic (Z-ENG prefix)
                is_english_topic = student.get('is_english_topic', False)
                
                if is_english_topic:
                    # Count English topics separately
                    english_student_count += 1
                else:
                    # Count Hungarian topics by category
                    category = student['topic_category']
                    hungarian_topic_categories[category] = hungarian_topic_categories.get(category, 0) + 1
        
        # Calculate required sessions
        total_sessions_needed = math.ceil(students_with_topics / 9) if students_with_topics > 0 else 0
        
        # Calculate planner table dimensions
        total_time_slots = len(self.time_slots) * len(self.weekdays) * len(self.rooms)
        
        return {
            'total_students': total_students,
            'students_with_topics': students_with_topics,
            'hungarian_topic_categories': hungarian_topic_categories,
            'english_student_count': english_student_count,
            'required_sessions': total_sessions_needed,
            'planner_dimensions': {
                'time_slots': len(self.time_slots),
                'weekdays': len(self.weekdays),
                'rooms': len(self.rooms),
                'total_slots': total_time_slots
            },
            'session_types_supported': self.all_session_types
        }
    
    def process_mkxlsx(self, fused_json_file: Optional[Path] = None,
                      output_excel_file: Optional[Path] = None) -> bool:
        """
        Complete MKXLSX process: load data and create Excel file
        
        Args:
            fused_json_file: Path to fused JSON file
            output_excel_file: Path for output Excel file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Load fused data
            if not self.load_fused_data(fused_json_file):
                self.logger.error("Failed to load fused data")
                return False
            
            # Create Excel file
            if not self.create_session_planning_excel(output_excel_file):
                self.logger.error("Failed to create Excel file")
                return False
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error in MKXLSX process: {e}")
            return False


def main():
    """Main function for testing the MKXLSX feature"""
    creator = SessionPlannerExcelCreator()
    
    # Process MKXLSX
    success = creator.process_mkxlsx()
    
    if success:
        # Print summary
        summary = creator.get_planning_summary()
        print("MKXLSX - Session Planning Excel Created!")
        print("=" * 50)
        print(f"📊 Planning Data Summary:")
        print(f"   • Total students: {summary['total_students']}")
        print(f"   • Students with topics: {summary['students_with_topics']}")
        print(f"   • Required sessions (9 per session): {summary['required_sessions']}")
        print()
        print(f"📚 Hungarian Topic Categories:")
        for category, count in summary['hungarian_topic_categories'].items():
            sessions_needed = math.ceil(count / 9)
            print(f"   • {category}: {count} students ({sessions_needed} sessions)")
        print()
        if summary['english_student_count'] > 0:
            sessions_needed = math.ceil(summary['english_student_count'] / 9)
            print(f"🌍 English Topics (ENG):")
            print(f"   • ENG: {summary['english_student_count']} students ({sessions_needed} sessions)")
            print()
        print(f"📅 Planner Table:")
        dims = summary['planner_dimensions']
        print(f"   • Time slots: {dims['time_slots']} ({', '.join(creator.time_slots)})")
        print(f"   • Weekdays: {dims['weekdays']} ({', '.join(creator.weekdays)})")
        print(f"   • Rooms: {dims['rooms']} ({', '.join(creator.rooms)})")
        print(f"   • Total planning slots: {dims['total_slots']}")
        print()
        print(f"🏷️  Supported Session Types:")
        print(f"   • Standard: {', '.join(creator.session_types)}")
        print(f"   • Special: {', '.join(creator.special_session_types)}")
        print()
        print("📁 Output: data/session_planner.xlsx")
        print("✅ MKXLSX completed successfully!")
    else:
        print("❌ MKXLSX failed!")


if __name__ == "__main__":
    main()
