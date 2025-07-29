"""
MKXLSX - Excel table creation for session planning (Refactored)
This module creates an Excel table using DLFUSION data for session schedule planning.
Refactored according to SOLID principles for better maintainability.
"""

import json
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math
import sys

sys.path.append(str(Path(__file__).parent.parent))
from shared import setup_logging
from excel_components import (
    ExcelStyler, ExcelTableBuilder, PlannerTableCreator, 
    StatisticsCalculator, StatisticsTableCreator
)


class SessionPlannerExcelCreatorRefactored:
    """Refactored creator for session planning Excel table"""
    
    def __init__(self):
        """Initialize the refactored Excel creator"""
        self.logger = setup_logging(__name__)
        self.fused_data = None
        self.workbook = None
        self.worksheet = None
        
        # Course codes to consider for statistics (from specification)
        self.STATISTICS_COURSE_CODES = {
            'BMEVIAUMT00', 'BMEVIAUMT10', 'BMEVIAUMT12', 'BMEVIAUM026', 
            'BMEVIAUAL01', 'BMEVIAUAL03', 'BMEVIAUAL04', 'BMEVIAUAL05', 
            'BMEVIAUML10', 'BMEVIAUML12', 'BMEVIAUML11', 'BMEVIAUML13', 
            'BMEVIAUM039'
        }
        
        # Initialize components
        self.styler = None
        self.table_builder = None
        self.planner_creator = None
        self.stats_calculator = None
        self.stats_creator = None
    
    def load_fused_data(self, json_file: Optional[Path] = None) -> bool:
        """
        Load fused data from JSON file
        
        Args:
            json_file: Path to fused JSON file (defaults to data/fused_student_data.json)
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if json_file is None:
                json_file = Path(__file__).parent.parent / "data" / "fused_student_data.json"
            
            if not json_file.exists():
                self.logger.error(f"Fused data file not found: {json_file}")
                return False
            
            self.logger.info(f"Loading fused data from: {json_file}")
            
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            self.fused_data = data
            self.logger.info(f"Loaded fused data with {len(data['students'])} students")
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading fused data: {e}")
            return False
    
    def create_session_planner_excel(self, output_file: Optional[Path] = None) -> bool:
        """
        Create the complete session planning Excel file
        
        Args:
            output_file: Path to output Excel file (defaults to data/session_planner.xlsx)
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if output_file is None:
                output_file = Path(__file__).parent.parent / "data" / "session_planner.xlsx"
            
            if not self.fused_data:
                self.logger.error("No fused data loaded. Call load_fused_data() first.")
                return False
            
            self.logger.info(f"Creating session planning Excel file: {output_file}")
            
            # Create workbook and worksheet
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Session Planner"
            
            # Initialize components
            self._initialize_components()
            
            # Add instructions and formatting
            self._add_instructions_and_formatting()
            
            # Create planner table
            current_row = self.planner_creator.create_planner_table(start_row=3)
            
            # Create statistics section
            current_row = self.stats_creator.create_statistics_section(current_row + 1, self.fused_data)
            
            # Save the workbook
            self.workbook.save(output_file)
            self.logger.info(f"Session planning Excel file created: {output_file}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating Excel file: {e}")
            return False
    
    def _initialize_components(self):
        """Initialize all component classes"""
        self.styler = ExcelStyler()
        self.table_builder = ExcelTableBuilder(self.worksheet, self.styler, self.logger)
        self.planner_creator = PlannerTableCreator(self.worksheet, self.styler, self.table_builder, self.logger)
        self.stats_calculator = StatisticsCalculator(self.logger, self.STATISTICS_COURSE_CODES)
        self.stats_creator = StatisticsTableCreator(
            self.worksheet, self.styler, self.table_builder, self.stats_calculator, self.logger
        )
    
    def _add_instructions_and_formatting(self):
        """Add instructions and basic formatting to the worksheet"""
        try:
            # Add instructions at the top
            instructions = ("Fill in the session types (SW, HW, ENG, GEN, ONLINE, SPARE, NONE, ROBONAUT, AI) "
                          "in the planner table below. The statistics will update automatically.")
            
            self.worksheet.cell(row=2, column=1, value=instructions)
            self.worksheet.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
            self.worksheet.cell(row=2, column=1).font = Font(italic=True, size=10)
            
            # Merge cells for instructions (row 2, columns A to P)
            self.worksheet.merge_cells('A2:P2')
            
            self.logger.info("Added instructions and formatting")
            
        except Exception as e:
            self.logger.error(f"Error adding instructions: {e}")
    
    def get_planning_summary(self) -> Dict[str, Any]:
        """
        Get a summary of the planning data
        
        Returns:
            Dictionary with planning statistics
        """
        if not self.fused_data:
            return {'error': 'No fused data available'}
        
        if not self.stats_calculator:
            self.stats_calculator = StatisticsCalculator(self.logger, self.STATISTICS_COURSE_CODES)
        
        # Calculate topic statistics
        topic_stats = self.stats_calculator.calculate_topic_statistics(self.fused_data)
        
        # Calculate planner table dimensions
        time_slots = 4  # 8:00-10:00, 10:00-12:00, 12:00-14:00, 14:00-16:00
        weekdays = 5   # Monday to Friday
        rooms = 3      # QB203, QBF14, QBF15
        total_time_slots = time_slots * weekdays * rooms
        
        return {
            'total_students': len(self.fused_data.get('students', [])),
            'students_with_topics': topic_stats['total_students'],
            'hungarian_topic_categories': topic_stats['hungarian_topic_categories'],
            'english_student_count': topic_stats['english_student_count'],
            'required_sessions': self.stats_calculator.calculate_required_sessions(topic_stats['total_students']),
            'planner_dimensions': {
                'time_slots': time_slots,
                'weekdays': weekdays,
                'rooms': rooms,
                'total_slots': total_time_slots
            }
        }


def main():
    """Main function for testing the refactored MKXLSX creator"""
    creator = SessionPlannerExcelCreatorRefactored()
    
    print("🚀 Testing MKXLSX feature...")
    print("=" * 50)
    
    # Load fused data
    if not creator.load_fused_data():
        print("❌ Failed to load fused data")
        return
    
    # Create Excel file
    if not creator.create_session_planner_excel():
        print("❌ Failed to create Excel file")
        return
    
    print("✅ MKXLSX completed successfully!")
    
    # Print summary
    summary = creator.get_planning_summary()
    
    print(f"📊 Session Planning Summary:")
    print(f"   • Total students: {summary['total_students']}")
    print(f"   • Students with topics: {summary['students_with_topics']}")
    print(f"   • Required sessions (9 per session): {summary['required_sessions']}")
    
    print(f"📅 Planner Table Structure:")
    dims = summary['planner_dimensions']
    print(f"   • Time slots: {dims['time_slots']} (8:00-10:00, 10:00-12:00, 12:00-14:00, 14:00-16:00)")
    print(f"   • Weekdays: {dims['weekdays']} (Monday, Tuesday, Wednesday, Thursday, Friday)")
    print(f"   • Rooms: {dims['rooms']} (QB203, QBF14, QBF15)")
    print(f"   • Total planning slots: {dims['total_slots']}")
    
    print(f"🎯 Supported Session Types:")
    print(f"   • Standard: SW, HW, ENG, GEN, ONLINE, SPARE, NONE")
    print(f"   • Special: ROBONAUT, AI")
    
    print(f"📁 Output file: ../data/session_planner.xlsx")
    print(f"💡 The Excel file contains:")
    print(f"   • Upper section: Planner table for entering session types")
    print(f"   • Lower section: Statistics with automatic Excel formulas")


if __name__ == "__main__":
    main()
