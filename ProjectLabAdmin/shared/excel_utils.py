"""
Excel utilities for reading and writing data
"""
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from typing import List, Dict, Any, Optional
import logging
from .config import EXCEL_OUTPUT_FILE, DEFAULT_SHEET_NAME

logger = logging.getLogger(__name__)


class ExcelHandler:
    """Handle Excel file operations"""
    
    def __init__(self, file_path: Optional[Path] = None):
        """
        Initialize Excel handler
        
        Args:
            file_path: Path to Excel file (defaults to config setting)
        """
        self.file_path = file_path or EXCEL_OUTPUT_FILE
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
    
    def create_excel_file(self, data: List[Dict[str, Any]], sheet_name: str = DEFAULT_SHEET_NAME) -> bool:
        """
        Create Excel file with data
        
        Args:
            data: List of dictionaries containing row data
            sheet_name: Name of the worksheet
            
        Returns:
            True if successful, False otherwise
        """
        try:
            df = pd.DataFrame(data)
            
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(self.file_path)
            logger.info(f"Excel file created successfully: {self.file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            return False
    
    def append_to_excel(self, data: List[Dict[str, Any]], sheet_name: str = DEFAULT_SHEET_NAME) -> bool:
        """
        Append data to existing Excel file
        
        Args:
            data: List of dictionaries containing row data
            sheet_name: Name of the worksheet
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not self.file_path.exists():
                return self.create_excel_file(data, sheet_name)
            
            # Load existing workbook
            wb = load_workbook(self.file_path)
            
            # Get or create worksheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
            
            # Convert data to DataFrame
            df = pd.DataFrame(data)
            
            # Find the next empty row
            next_row = ws.max_row + 1
            
            # Append data
            for r in dataframe_to_rows(df, index=False, header=False):
                ws.append(r)
            
            wb.save(self.file_path)
            logger.info(f"Data appended to Excel file: {self.file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error appending to Excel file: {e}")
            return False
    
    def read_excel_data(self, sheet_name: str = DEFAULT_SHEET_NAME) -> Optional[List[Dict[str, Any]]]:
        """
        Read data from Excel file
        
        Args:
            sheet_name: Name of the worksheet to read
            
        Returns:
            List of dictionaries containing row data, or None if failed
        """
        try:
            if not self.file_path.exists():
                logger.error(f"Excel file not found: {self.file_path}")
                return None
            
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
            data = df.to_dict('records')
            logger.info(f"Successfully read {len(data)} rows from Excel file")
            return data
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            return None
    
    def get_sheet_names(self) -> List[str]:
        """
        Get list of sheet names in the Excel file
        
        Returns:
            List of sheet names
        """
        try:
            if not self.file_path.exists():
                return []
            
            wb = load_workbook(self.file_path)
            return wb.sheetnames
            
        except Exception as e:
            logger.error(f"Error getting sheet names: {e}")
            return []
